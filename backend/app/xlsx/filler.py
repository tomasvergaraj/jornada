"""Rellena las planillas oficiales en .xlsx con openpyxl, preservando logos,
estilos, celdas combinadas, formatos de hora y fórmulas. Las horas se escriben
como fracción de día con su formato (h:mm / [h]:mm), igual que Excel.

openpyxl no sabe round-tripear formas vectoriales (los recuadros/casillas donde
va la X, conectores, cuadros de texto): al guardar regenera el dibujo y las
descarta. Por eso, tras rellenar las celdas, restauramos el dibujo original
(xl/drawings/* + xl/media/*) tal cual desde la plantilla — ver _merge_dibujos."""
from __future__ import annotations
import io
import re
import zipfile
from copy import copy
from collections import Counter
from datetime import datetime
from openpyxl import load_workbook
from .forms import FORMS, MESES

# Partes que openpyxl no preserva fielmente y que reponemos desde la plantilla.
# (drawings incluye también vmlDrawing y los _rels del dibujo; media son logos/
#  imágenes, incl. el .wdp HD-Photo que openpyxl elimina.)
_PRESERVE_PREFIXES = ("xl/drawings/", "xl/media/")


def _frac(mins: int | None):
    return None if mins is None else round(mins / 1440.0, 10)


def _merge_content_types(filled_ct: bytes, orig_ct: bytes, final_names: set[str]) -> bytes:
    """Fusiona el [Content_Types].xml de openpyxl con el de la plantilla: añade
    los <Default> de extensiones que falten (p. ej. wdp) y los <Override> de las
    partes restauradas que falten, sin arrastrar overrides de partes ausentes."""
    f = filled_ct.decode("utf-8")
    o = orig_ct.decode("utf-8")
    have_ext = set(re.findall(r'<Default\b[^>]*?Extension="([^"]+)"', f))
    have_part = set(re.findall(r'<Override\b[^>]*?PartName="([^"]+)"', f))
    adds: list[str] = []
    for m in re.finditer(r'<Default\b[^>]*?Extension="([^"]+)"[^>]*?/>', o):
        if m.group(1) not in have_ext:
            adds.append(m.group(0)); have_ext.add(m.group(1))
    for m in re.finditer(r'<Override\b[^>]*?PartName="([^"]+)"[^>]*?/>', o):
        part = m.group(1)
        if part not in have_part and part.lstrip("/") in final_names:
            adds.append(m.group(0)); have_part.add(part)
    if adds:
        f = f.replace("</Types>", "".join(adds) + "</Types>")
    return f.encode("utf-8")


def _merge_dibujos(template_path: str, filled_bytes: bytes) -> bytes:
    """Toma el .xlsx que produjo openpyxl (celdas correctas) y le repone el
    dibujo original de la plantilla (formas, casillas, conectores, logos) que
    openpyxl había descartado. La hoja ya referencia xl/drawings/drawing1.xml
    por ruta, así que basta reponer el contenido y fusionar el content-types."""
    with zipfile.ZipFile(template_path) as orig, \
         zipfile.ZipFile(io.BytesIO(filled_bytes)) as filled:
        preserve = [n for n in orig.namelist() if n.startswith(_PRESERVE_PREFIXES)]
        # nombres del paquete final = lo que conserva openpyxl (menos lo repuesto)
        # más las partes originales restauradas.
        kept = [n for n in filled.namelist()
                if not n.startswith(_PRESERVE_PREFIXES) and n != "[Content_Types].xml"]
        final_names = set(kept) | set(preserve) | {"[Content_Types].xml"}

        out = io.BytesIO()
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
            for n in kept:
                z.writestr(n, filled.read(n))
            for n in preserve:                      # dibujo + medios originales
                z.writestr(n, orig.read(n))
            z.writestr("[Content_Types].xml", _merge_content_types(
                filled.read("[Content_Types].xml"),
                orig.read("[Content_Types].xml"), final_names))
        return out.getvalue()


def _ident(ws, cfg) -> dict:
    def g(addr):
        v = ws[addr].value
        return "" if v is None else str(v).strip()
    i = cfg["ident"]
    return {k: g(a) for k, a in i.items()}


def fill(form_key: str, template_path: str, jornadas: list[dict], mes: int, anio: int,
         perfil: dict | None = None) -> tuple[bytes, dict]:
    cfg = FORMS[form_key]
    wb = load_workbook(template_path)
    ws = wb.active

    # identidad: si llega un perfil, se escribe en la planilla; si no, se
    # conserva (y devuelve) la que ya trae la plantilla.
    if perfil:
        for campo, addr in cfg["ident"].items():
            val = perfil.get(campo)
            if val:
                ws[addr] = val
    ident = _ident(ws, cfg)

    # limpiar filas de días
    lo, hi = cfg["day_rows"]
    for r in range(lo, hi + 1):
        for col in ("C", "D", "E", "F", "G"):
            ws[f"{col}{r}"] = None

    # número del día (columna B "FECHA"): se escribe como valor literal. La
    # plantilla lo trae como cadena de fórmulas (=B19+1 …) que openpyxl guarda
    # sin valor en caché, y varios visores la muestran en blanco. Con literales
    # el número siempre aparece, sin depender del recálculo.
    for r in range(lo, hi + 1):
        ws[f"B{r}"] = r - cfg["day_offset"]

    # fuente estándar de la columna de justificación: la mayoritaria entre las
    # filas de días. Algunas plantillas traen una celda suelta con otra fuente
    # (p. ej. el día 29 en comp venía en Calibri/negrita); al escribir la
    # justificación se normaliza para que todos los días salgan igual.
    _cnt: Counter = Counter()
    _ref: dict = {}
    for r in range(lo, hi + 1):
        fo = ws[f"C{r}"].font
        k = (fo.name, fo.size, fo.bold, fo.italic)
        _cnt[k] += 1
        _ref.setdefault(k, fo)
    fuente_just = _ref[_cnt.most_common(1)[0][0]]

    # mes / año
    if cfg["mes"]["tipo"] == "fecha":
        c = ws[cfg["mes"]["celda"]]
        c.value = datetime(anio, mes, 1)
        c.number_format = cfg["mes"]["fmt"]
    else:
        ws[cfg["mes"]["celda_mes"]] = MESES[mes - 1]
        ws[cfg["mes"]["celda_anio"]] = anio

    tot_d = tot_f = 0
    for j in jornadas:
        d = int(j["fecha"][8:10])
        rr = cfg["day_offset"] + d
        if j.get("justificacion"):
            cC = ws[f"C{rr}"]
            cC.value = j["justificacion"].strip()
            cC.font = copy(fuente_just)
        if j.get("ingreso_min") is not None:
            c = ws[f"D{rr}"]; c.value = _frac(j["ingreso_min"]); c.number_format = "h:mm"
        if j.get("salida_min") is not None:
            c = ws[f"E{rr}"]; c.value = _frac(j["salida_min"]); c.number_format = "h:mm"
        if j.get("diurna_min", 0) > 0:
            c = ws[f"F{rr}"]; c.value = _frac(j["diurna_min"]); c.number_format = "[h]:mm"; tot_d += j["diurna_min"]
        if j.get("fest_min", 0) > 0:
            c = ws[f"G{rr}"]; c.value = _frac(j["fest_min"]); c.number_format = "[h]:mm"; tot_f += j["fest_min"]

    # totales (fórmula SUM; Excel/LibreOffice recalculan al abrir)
    ws[cfg["total_diurna"]] = f"=SUM({cfg['rango_diurna']})"
    ws[cfg["total_diurna"]].number_format = "[h]:mm"
    ws[cfg["total_fest"]] = f"=SUM({cfg['rango_fest']})"
    ws[cfg["total_fest"]].number_format = "[h]:mm"

    buf = io.BytesIO()
    wb.save(buf)
    data = _merge_dibujos(template_path, buf.getvalue())
    return data, {"ident": ident, "total_diurna_min": tot_d, "total_fest_min": tot_f}
