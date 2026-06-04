from fastapi import APIRouter
import os
from openpyxl import load_workbook
from ..schemas import Perfil
from ..xlsx.forms import FORMS
from ..config import settings

router = APIRouter()


def _perfil_desde_plantilla() -> Perfil | None:
    """Lee la identidad escrita en la plantilla comp (nombre, RUT, etc.)."""
    p = os.path.join(settings.templates_dir, FORMS["comp"]["archivo"])
    if not os.path.exists(p):
        return None
    try:
        ws = load_workbook(p).active
        i = FORMS["comp"]["ident"]
        g = lambda a: (str(ws[a].value).strip() if ws[a].value is not None else "")
        return Perfil(nombre=g(i["nombre"]), rut=g(i["rut"]), cargo=g(i["cargo"]),
                      grado=g(i["grado"]), servicio=g(i["servicio"]))
    except Exception:
        return None


@router.get("/profile", response_model=Perfil)
def get_profile():
    """Sugerencia de identidad tomada de la plantilla. El perfil real de cada
    persona vive en su navegador (localStorage), no en el servidor; este
    endpoint solo sirve para ofrecer un autocompletado inicial."""
    return _perfil_desde_plantilla() or Perfil()
